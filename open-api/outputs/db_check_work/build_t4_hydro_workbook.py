# -*- coding: utf-8 -*-
"""실측가뭄 T4 정형데이터 — 수력발전량 미비점 점검 워크북 생성.

읽기 전용 DB 조회 결과 + 파일 조사 결과를 11개 시트 엑셀로 정리한다.
DB/파일을 변경하지 않는다.
"""
import os
from datetime import date

from dotenv import load_dotenv

load_dotenv("/Users/jeongseok/Desktop/workspace_model/dam-hydropower-prediction-model/.env")

import pandas as pd
from sqlalchemy import create_engine
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/Users/jeongseok/Desktop/workspace_intelliJ/icuh-platform-api/outputs/db_check_work"
OUT_PATH = f"{OUT_DIR}/실측가뭄_T4_수력발전량_미비점점검_v1.xlsx"
SURVEY_MD = f"{OUT_DIR}/실측가뭄_사례검증_DB현황조사_20260805.md"
REPO = "/Users/jeongseok/Desktop/workspace_model/dam-hydropower-prediction-model"
CHECK_DATE = "2026-08-06"

# 운영 DB(AWS RDS icuh-rds) — EC2 SSH 터널 13306 경유. 터널이 없으면 RDS 항목은 '미확인'으로 표기된다.
RDS_URL = "mysql+pymysql://dasom:dasom123@127.0.0.1:13306/ACTUAL_DRGHT?charset=utf8mb4"

SOUTH, GN = "CASE_SOUTH_22_23", "CASE_GANGNEUNG_2025"
SOUTH_RANGE, GN_RANGE = (202201, 202312), (202501, 202512)

engine = create_engine(os.environ["DB_URL"])


# ---------------------------------------------------------------- 데이터 적재
def load_frames():
    daily = pd.read_sql(
        "SELECT dam_name, dam_code, generation_date, planned_mwh, actual_mwh FROM dam_daily_generation", engine)
    daily["generation_date"] = pd.to_datetime(daily.generation_date)
    for c in ("planned_mwh", "actual_mwh"):
        daily[c] = pd.to_numeric(daily[c])

    gen = pd.read_sql(
        "SELECT dam_name, dam_code, CAST(year AS UNSIGNED)*100+CAST(month AS UNSIGNED) ymi,"
        " planned_mwh, actual_mwh FROM dam_monthly_generation", engine)
    res = pd.read_sql(
        "SELECT dam_name, CAST(year AS UNSIGNED)*100+CAST(month AS UNSIGNED) ymi,"
        " water_level_elm, water_storage_mcm FROM dam_monthly_reservoir_status", engine)
    pred = pd.read_sql(
        "SELECT dam_name, CAST(year AS UNSIGNED)*100+CAST(month AS UNSIGNED) ymi,"
        " predicted_power_generation_lower_bound pl, predicted_power_generation_upper_bound pu,"
        " predicted_water_storage_lower_bound sl, predicted_water_storage_upper_bound su"
        " FROM dam_monthly_predictions", engine)
    comp = pd.read_sql(
        "SELECT dam_name, CAST(year AS UNSIGNED)*100+CAST(month AS UNSIGNED) ymi FROM dam_monthly_comparison", engine)
    for df, cols in ((gen, ["planned_mwh", "actual_mwh"]), (res, ["water_level_elm", "water_storage_mcm"]),
                     (pred, ["pl", "pu", "sl", "su"])):
        for c in cols:
            df[c] = pd.to_numeric(df[c])
    return daily, gen, res, pred, comp


def monthly_from_cumulative(daily):
    """연누계 원자료를 월말 누계 차분으로 '실제 월별 발전량'으로 환산."""
    out = []
    for dam, g in daily.groupby("dam_name"):
        g = g.sort_values("generation_date").copy()
        g["ym"] = g.generation_date.dt.to_period("M")
        eom = g.groupby("ym").agg(last_day=("generation_date", "max"), n_days=("generation_date", "size"),
                                  cum_a=("actual_mwh", "last"), cum_p=("planned_mwh", "last"))
        eom["pa"], eom["pp"] = eom.cum_a.shift(1), eom.cum_p.shift(1)
        eom.loc[eom.index.month == 1, ["pa", "pp"]] = 0.0
        eom["gen_actual_mwh"] = eom.cum_a - eom.pa
        eom["gen_planned_mwh"] = eom.cum_p - eom.pp
        eom["dam_name"] = dam
        out.append(eom.reset_index())
    m = pd.concat(out, ignore_index=True)
    m["ymi"] = m.ym.dt.year * 100 + m.ym.dt.month
    return m


daily, gen, res, pred, comp = load_frames()
real = monthly_from_cumulative(daily)


# ------------------------------------------------------- 운영 RDS 조회(있으면)
def load_rds():
    """운영 RDS에서 WT_POWER 2종·발전소 운영 4종·중복 현황을 가져온다. 실패 시 None."""
    try:
        r = create_engine(RDS_URL, connect_args={"connect_timeout": 8})
        with r.connect() as c:
            c.execute(__import__("sqlalchemy").text("SELECT 1"))
    except Exception as ex:
        print("[경고] RDS 미접속 —", type(ex).__name__, "→ RDS 항목은 '미확인'으로 기록")
        return None
    out = {}
    out["info"] = pd.read_sql("SELECT STN_ID, STN_NAME, ADDRESS FROM WT_POWER_INFO ORDER BY STN_ID", r)
    wp = pd.read_sql("SELECT DISTINCT STN_ID, STN_NAME, ADDRESS, CRT_YMD, POWER FROM WT_POWER", r)
    wp["d"] = pd.to_datetime(wp.CRT_YMD, errors="coerce")
    out["wp"] = wp
    out["wp_total"] = int(pd.read_sql("SELECT COUNT(*) n FROM WT_POWER", r).n[0])
    ops = {}
    for dam, t in [("소양강", "drought_impact_soyangriver_dam_operation"),
                   ("충주", "drought_impact_chungju_dam_operation"),
                   ("대청", "drought_impact_daecheong_dam_operation"),
                   ("합천", "drought_impact_hapcheon_dam_operation")]:
        d = pd.read_sql(f"SELECT DISTINCT elcpwstnCd, elcpwstnNm, stdrDe, r2, s2, t2 FROM `{t}`", r)
        d["stdrDe"] = pd.to_datetime(d.stdrDe)
        d["ymi"] = d.stdrDe.dt.year * 100 + d.stdrDe.dt.month
        d["_table"] = t
        d["_total"] = int(pd.read_sql(f"SELECT COUNT(*) n FROM `{t}`", r).n[0])
        ops[dam] = d
    out["ops"] = ops
    dup = {}
    for t, key in [("dam_daily_generation", "dam_code, generation_date"),
                   ("dam_monthly_generation", "dam_code, year, month"),
                   ("dam_monthly_reservoir_status", "dam_code, year, month"),
                   ("dam_monthly_predictions", "dam_code, year, month"),
                   ("dam_monthly_comparison", "dam_code, year, month")]:
        tot = int(pd.read_sql(f"SELECT COUNT(*) n FROM `{t}`", r).n[0])
        uk = int(pd.read_sql(f"SELECT COUNT(*) n FROM (SELECT DISTINCT {key} FROM `{t}`) x", r).n[0])
        n_idx = len(pd.read_sql(f"SHOW INDEX FROM `{t}`", r))
        dup[t] = dict(total=tot, unique=uk, excess=tot - uk, indexes=n_idx)
    out["dup"] = dup
    return out


RDS = load_rds()
RDS_OK = RDS is not None


def rds_note(fn, fallback="RDS 미접속으로 미확인"):
    try:
        return fn() if RDS_OK else fallback
    except Exception:
        return fallback

DAMS = ["소양강", "충주", "대청", "합천"]
DAM_META = {
    "소양강": dict(code="1012110", sido="강원특별자치도", sgg="춘천시", river="한강수계(북한강)",
                stn="소양강수력발전소", cap=2900, eff=1900),
    "충주":  dict(code="1003110", sido="충청북도", sgg="충주시", river="한강수계(남한강)",
                stn="충주수력발전소", cap=2750, eff=1789),
    "대청":  dict(code="3008110", sido="대전광역시", sgg="대덕구", river="금강수계",
                stn="대청수력발전소", cap=1490, eff=790),
    "합천":  dict(code="2015110", sido="경상남도", sgg="합천군", river="낙동강수계(황강)",
                stn="합천수력발전소", cap=790, eff=560),
}

# 사례 × 댐 공간판정
VERDICT = {
    (SOUTH, "합천"): ("직접 검증자료", "사용",
                    "댐 소재지가 사례 핵심지역인 경남 합천군과 행정구역이 일치한다(황강·낙동강수계)."),
    (SOUTH, "대청"): ("적용 부적합", "미사용",
                    "금강수계(대전·충북)로 사례지역(전남 고흥·경남 합천)과 유역·행정구역이 모두 다르다."),
    (SOUTH, "충주"): ("적용 부적합", "미사용",
                    "한강수계(충북 충주)로 사례지역과 유역·행정구역이 모두 다르다."),
    (SOUTH, "소양강"): ("적용 부적합", "미사용",
                     "한강수계(강원 춘천)로 사례지역과 유역·행정구역이 모두 다르다."),
    (GN, "소양강"): ("보조자료", "참고만",
                  "강원권 최대 다목적댐이나 태백산맥 서편 영서(북한강)로 강릉(영동 동해안 소하천)과 유역·강수패턴이 다르다. "
                  "강원 광역 수자원 상황 서술용 보조지표로만 사용."),
    (GN, "충주"): ("적용 부적합", "미사용", "한강수계(충북)로 강릉 동해안 권역과 무관하다."),
    (GN, "대청"): ("적용 부적합", "미사용", "금강수계로 강릉 동해안 권역과 무관하다."),
    (GN, "합천"): ("적용 부적합", "미사용", "낙동강수계(경남)로 강릉 동해안 권역과 무관하다."),
}
CASE_REGION = {SOUTH: "전남 고흥군 / 경남 합천군", GN: "강원 강릉시(대관령·평창 관련지역)"}


def ym_list(lo, hi):
    return [y * 100 + m for y in range(lo // 100, hi // 100 + 1) for m in range(1, 13) if lo <= y * 100 + m <= hi]


SOUTH_YMS, GN_YMS = ym_list(*SOUTH_RANGE), ym_list(*GN_RANGE)

# ------------------------------------------------------------ 시트 1. README
readme = [
    ["실측가뭄 T4 정형데이터 — 수력발전량 미비점 점검", ""],
    ["", ""],
    ["점검일", CHECK_DATE],
    ["점검자", "이정석 (인프라재난관리진흥원)"],
    ["점검 목적", "T4 정형데이터 구축 단계에서 '수력발전량'이 두 가뭄 사례 검증자료로 사용 가능한지 판정"],
    ["", ""],
    ["사례 1", "2022~2023 남부지방 가뭄 / 핵심지역: 전남 고흥군, 경남 합천군 / 대상기간 2022-01~2023-12"],
    ["사례 2", "2025 강릉지역 가뭄 / 핵심지역: 강원 강릉시(대관령·평창 관련) / 대상기간 2025-01~2025-12"],
    ["", ""],
    ["점검 대상 DB(1)", "MySQL ACTUAL_DRGHT (로컬 127.0.0.1:3306) — 읽기 전용 SELECT만 수행, 원본 무변경"],
    ["점검 대상 DB(2)", "AWS RDS icuh-rds / ACTUAL_DRGHT (EC2 3.37.105.126 SSH 터널 경유) — "
     + ("본 점검에서 직접 조회 완료. 읽기 전용" if RDS_OK else "본 점검에서 미접속")],
    ["점검 대상 코드", f"{REPO} (dam-hydropower-prediction-model 파이프라인)"],
    ["참고 선행조사", f"{SURVEY_MD} (2026-08-05, AWS RDS icuh-rds 대상)"],
    ["", ""],
    ["■ 핵심 결론 (요약)", ""],
    ["1", "수력발전량 데이터는 ACTUAL_DRGHT의 dam_* 5종 테이블에 존재하며 2022·2023·2025년 모두 결측 없이 보유한다."],
    ["2", "dam_* 대상 댐은 소양강·충주·대청·합천 4개뿐이다. 고흥군·강릉시에 대응하는 댐은 없다."],
    ["3", "합천댐만 사례지역(경남 합천군)과 행정구역이 일치하여 '직접 검증자료'로 사용 가능하다."],
    ["4", "소양강댐은 강원권이지만 영서(북한강)로 강릉(영동)과 유역이 달라 '보조자료'로만 제한 사용한다."],
    ["5", "충주·대청댐은 두 사례 모두 '적용 부적합'이다."],
    ["6", "고흥군은 대응 댐이 없어 수력발전량 지표를 만들 수 없다(미보유). WT_POWER의 보성강(전남 보성군)·"
          "칠보(전북 정읍시)는 고흥군이 아니고 보유 구간도 2023-07~12뿐이다."],
    ["6-1", "[신규 확인] 강릉 사례와 공간적으로 직접 연결되는 발전소는 실재한다 — WT_POWER STN_ID 3210, "
            "주소 '강원도 강릉시 성산면 오봉리'(도암댐 유역변경 방류를 받는 강릉수력). "
            "그러나 ① 보유 구간이 2023-07-26~12-26(92일)뿐이고 ② 그 구간의 POWER 값이 92행 전부 0이며 "
            "③ 2025년 자료가 0행이라, 강릉 사례에는 적용할 수 없다."],
    ["6-2", "[신규 확인] 선행조사의 미해소 항목이던 'WT_POWER와 WT_POWER_INFO의 발전소명 불일치'는 해소되었다. "
            "STN_ID로 조인하면 10개소가 1:1로 매칭되며 주소도 9/10 일치한다. 명칭만 3건 다르다"
            "(3130 안흥=강림, 3180 섬진강=칠보, 3210 강릉=도암). 매칭 키는 STN_ID를 쓰면 된다."],
    ["7", "[중대] actual_mwh/planned_mwh 원자료는 '연간 누계' 값이다. 월별 테이블 값은 누계의 월평균이므로 "
          "그대로 '월 발전량'으로 해석하면 안 된다. 월말 누계 차분으로 재환산해야 하며 본 워크북에 환산값을 함께 제시했다."],
    ["8", "[중대] dam_monthly_predictions의 2022-01~2024-10 구간은 모델 예측이 아니라 실적 ±80MWh(저수량 ±40) "
          "고정밴드다(2024-11부터 실제 모델 산출물). 남부 사례기간(2022~2023)의 '예측 대비 실적 오차' 검증은 불가능하다."],
    ["9", "2025년 예측값은 실제 모델 산출물이며 발전량 구간적중 70.8%, 저수량 95.8%로 사용 가능하다(구간폭 과대 구간 존재)."],
    ["10", "저수율(%) 컬럼은 DB에 없다. 저수량(백만㎥)만 보유하므로 댐 제원(총저수용량) 확보 후 산정해야 한다."],
    ["11", "[신규 확인] 운영 RDS의 dam_* 4종에 인덱스가 0개다(PRIMARY KEY 없음). 그 결과 적재가 멱등하지 않아 "
           "완전 동일한 행이 반복 저장되어 있다(dam_daily_generation 20,892행 중 고유 7,660 — 초과 13,232행, "
           "id 값까지 중복). 값 충돌은 없고 로컬 값과 일치하므로 SELECT DISTINCT로 안전하게 제거할 수 있다. "
           "T4 집계 시 반드시 중복 제거 후 사용할 것."],
    ["12", "[신규 확인] RDS에만 있는 drought_impact_{soyangriver,chungju,daecheong,hapcheon}_dam_operation 4종은 "
           "2024-01~2025-10 월별 자료다. 컬럼 주석이 실제 값과 어긋난다 — 주석상 '저수량(ton)'인 r2가 "
           "4개 댐 모두에서 월 발전량과 상관 +0.996~+0.999(중앙비율 0.995~1.000)로, 사실상 월 발전량(MWh)이다. "
           "주석상 '발전량(Mwh)'인 s2는 발전량·저수량 어느 쪽과도 맞지 않아 의미 미상이다. 사용 전 출처 확인 필요."],
    ["", ""],
    ["■ T4 수력발전량 항목 종합 상태", "진행 중 (합천댐 한정 사용 가능 / 연누계 환산·예측유형 구분·중복 제거 작업 필요)"],
    ["", ""],
    ["■ 표현 원칙 (준수)", ""],
    ["금지", "'가뭄 때문에 발전량이 감소했다' / '발전량 감소의 원인은 가뭄이다' — 인과 단정 표현"],
    ["사용", "'같은 시기에 발전량 또는 저수량 변화가 관측되었다'"],
    ["사용", "'수력발전량은 사례지역과 공간적으로 직접 연결되지 않아 보조자료로만 사용한다'"],
    ["사용", "'공간 연관성이 낮아 본 검증표에서는 제외한다'"],
    ["", ""],
    ["■ 시트 구성", ""],
    ["1. README", "본 시트"],
    ["2. 테이블_파일목록", "수력발전 관련 테이블·파일 구조와 기간·행수"],
    ["3. 댐발전소_목록", "대상 댐/발전소 목록과 위치·수계"],
    ["4. 실제발전량_점검", "실제 발전량 존재·단위·연도별 커버리지·결측"],
    ["5. 계획예측발전량_점검", "계획/예측 발전량 존재 여부와 비교 가능성"],
    ["6. 저수량저수율_점검", "저수량·저수위 보유 현황과 저수율 산정 가능성"],
    ["7. 공간적절성_판정", "사례 × 댐 공간판정(직접/보조/부적합/확인필요)과 근거"],
    ["8. 월별통합가능성", "요청 양식의 월별 검증표를 실제 값으로 채운 결과"],
    ["9. 미비점판정", "항목별 상태·근거·미비점·필요작업 판정표"],
    ["10. 추출쿼리_파일경로", "본 점검에 사용한 SQL과 파일 경로"],
    ["11. 품질점검", "결측·중복·이상값·의미불일치 등 품질 확인 결과"],
]

# ------------------------------------------------- 시트 2. 테이블_파일목록
def rng(df, col="ymi"):
    return f"{df[col].min()} ~ {df[col].max()}"


tbl_rows = [
    dict(구분="DB테이블", 위치="ACTUAL_DRGHT", 테이블_파일명="dam_daily_generation",
         주요컬럼="id, dam_name, dam_code, generation_date, planned_mwh, actual_mwh, created_at, updated_at",
         날짜컬럼="generation_date (date)", 댐발전소명컬럼="dam_name / dam_code", 위치컬럼="없음",
         실제발전량컬럼="actual_mwh (decimal 15,2)", 계획발전량컬럼="planned_mwh (decimal 15,2)",
         예측발전량컬럼="없음", 저수량저수율컬럼="없음", 시간단위="일",
         데이터기간=f"{daily.generation_date.min():%Y-%m-%d} ~ {daily.generation_date.max():%Y-%m-%d}",
         행수=len(daily), 비고="[중요] 값이 연간 누계. 매년 1월 1일은 전년 누계 총량(이상치)"),
    dict(구분="DB테이블", 위치="ACTUAL_DRGHT", 테이블_파일명="dam_monthly_generation",
         주요컬럼="id, dam_name, dam_code, year, month, planned_mwh, actual_mwh",
         날짜컬럼="year, month (varchar, 무패딩)", 댐발전소명컬럼="dam_name / dam_code", 위치컬럼="없음",
         실제발전량컬럼="actual_mwh", 계획발전량컬럼="planned_mwh", 예측발전량컬럼="없음",
         저수량저수율컬럼="없음", 시간단위="월", 데이터기간=rng(gen), 행수=len(gen),
         비고="일별 '연누계' 값의 월평균. 월 발전량 아님"),
    dict(구분="DB테이블", 위치="ACTUAL_DRGHT", 테이블_파일명="dam_monthly_reservoir_status",
         주요컬럼="id, dam_name, dam_code, year, month, water_level_elm, water_storage_mcm",
         날짜컬럼="year, month", 댐발전소명컬럼="dam_name / dam_code", 위치컬럼="없음",
         실제발전량컬럼="없음", 계획발전량컬럼="없음", 예측발전량컬럼="없음",
         저수량저수율컬럼="water_storage_mcm(백만㎥), water_level_elm(EL.m) — 저수율(%) 없음",
         시간단위="월", 데이터기간=rng(res), 행수=len(res), 비고="일별 저수량(현재)의 월평균"),
    dict(구분="DB테이블", 위치="ACTUAL_DRGHT", 테이블_파일명="dam_monthly_predictions",
         주요컬럼="id, dam_name, dam_code, year, month, predicted_power_generation_lower/upper_bound,"
                " predicted_water_storage_lower/upper_bound",
         날짜컬럼="year, month", 댐발전소명컬럼="dam_name / dam_code", 위치컬럼="없음",
         실제발전량컬럼="없음", 계획발전량컬럼="없음",
         예측발전량컬럼="predicted_power_generation_lower_bound / upper_bound (구간, 점추정 없음)",
         저수량저수율컬럼="predicted_water_storage_lower/upper_bound", 시간단위="월",
         데이터기간=rng(pred), 행수=len(pred),
         비고="2022-01~2024-10은 실적 ±80 고정밴드(모델 아님), 2024-11~는 Quantile GBR+conformal 산출물"),
    dict(구분="DB테이블", 위치="ACTUAL_DRGHT", 테이블_파일명="dam_monthly_comparison",
         주요컬럼="전년동월/전월 대비 발전량 amount·status·rate·color, 전월 대비 저수량 amount·status·rate·color",
         날짜컬럼="year, month", 댐발전소명컬럼="dam_name / dam_code", 위치컬럼="없음",
         실제발전량컬럼="없음(증감액)", 계획발전량컬럼="없음", 예측발전량컬럼="없음",
         저수량저수율컬럼="average_water_storage_last_month_amount", 시간단위="월",
         데이터기간=rng(comp), 행수=len(comp),
         비고="2022년 전체 없음. 2024-11·2024-12 결측. 화면 표출용 사전계산값"),
    dict(구분="DB테이블(RDS 전용)", 위치="AWS RDS ACTUAL_DRGHT", 테이블_파일명="WT_POWER",
         주요컬럼="IDX, STN_ID, STN_NAME, ADDRESS, CRT_YMD, POWER", 날짜컬럼="CRT_YMD (varchar 'YYYY-MM-DD')",
         댐발전소명컬럼="STN_ID / STN_NAME", 위치컬럼="ADDRESS (수력 테이블 중 유일하게 주소 보유)",
         실제발전량컬럼="POWER (int, 단위 미표기 — 확인 필요)", 계획발전량컬럼="없음", 예측발전량컬럼="없음",
         저수량저수율컬럼="없음", 시간단위="일",
         데이터기간=rds_note(lambda: f"{RDS['wp'].d.min():%Y-%m-%d} ~ {RDS['wp'].d.max():%Y-%m-%d} "
                                 f"(고유 {RDS['wp'].d.nunique()}일 / 기간내 결측 62일)"),
         행수=rds_note(lambda: f"{RDS['wp_total']} (중복 제거 시 {len(RDS['wp'])})"),
         비고="발전소 10개소. 로컬 DB에는 없음. 두 사례기간(2022~2023 상반기, 2025)을 모두 벗어남"),
    dict(구분="DB테이블(RDS 전용)", 위치="AWS RDS ACTUAL_DRGHT", 테이블_파일명="WT_POWER_INFO",
         주요컬럼="STN_ID, STN_NAME, ADDRESS", 날짜컬럼="없음(마스터)",
         댐발전소명컬럼="STN_ID / STN_NAME", 위치컬럼="ADDRESS",
         실제발전량컬럼="없음", 계획발전량컬럼="없음", 예측발전량컬럼="없음", 저수량저수율컬럼="없음",
         시간단위="-", 데이터기간="-",
         행수=rds_note(lambda: len(RDS["info"])),
         비고="발전소 마스터. WT_POWER와 STN_ID로 1:1 매칭(명칭만 3건 상이). 강릉(3210) 포함"),
    dict(구분="DB테이블(RDS 전용)", 위치="AWS RDS ACTUAL_DRGHT",
         테이블_파일명="drought_impact_{soyangriver,chungju,daecheong,hapcheon}_dam_operation",
         주요컬럼="elcpwstnCd(발전소코드), elcpwstnClsNm, elcpwstnNm(발전소명), r2, s2, t2, stdrDe",
         날짜컬럼="stdrDe (date, 매월 1일)", 댐발전소명컬럼="elcpwstnCd / elcpwstnNm",
         위치컬럼="없음", 실제발전량컬럼="r2 (주석은 '저수량(ton)'이나 실제로는 월 발전량 MWh로 확인)",
         계획발전량컬럼="없음", 예측발전량컬럼="없음",
         저수량저수율컬럼="없음 (s2가 '발전량(Mwh)' 주석이나 의미 미상)", 시간단위="월",
         데이터기간=rds_note(lambda: f"{RDS['ops']['합천'].ymi.min()} ~ {RDS['ops']['합천'].ymi.max()}"),
         행수=rds_note(lambda: " / ".join(f"{k} {int(v['_total'].iloc[0])}(고유 {len(v)})"
                                        for k, v in RDS["ops"].items())),
         비고="[중요] 4개 댐의 월 발전량을 '연누계 환산 없이' 바로 제공. 단 2024-01부터라 남부 사례기간 없음. "
             "컬럼 주석 오류 있음 — 사용 전 출처 확인 필요"),
    dict(구분="DB테이블(빈껍데기)", 위치="actual_drght_v3 / v4", 테이블_파일명=
         "hydro_generation_daily_raw, hydro_generation_monthly_summary, hydro_reservoir_monthly_status,"
         " hydro_operation_forecast_monthly, hydro_power_station 등",
         주요컬럼="차세대 스키마(설계만)", 날짜컬럼="-", 댐발전소명컬럼="plant_code / plant_name",
         위치컬럼="region_code", 실제발전량컬럼="-", 계획발전량컬럼="-", 예측발전량컬럼="-",
         저수량저수율컬럼="-", 시간단위="-", 데이터기간="없음", 행수=0,
         비고="전 테이블 0행. T4 사용 불가"),
    dict(구분="DB테이블(샘플)", 위치="actual_drght_v2", 테이블_파일명="hydro_power_monthly / hydro_power_station",
         주요컬럼="plant_code, base_month, reservoir_volume, power_generation, discharge_volume",
         날짜컬럼="base_month(YYYYMM)", 댐발전소명컬럼="plant_name(충주수력)", 위치컬럼="region_code(43130 충주시)",
         실제발전량컬럼="power_generation", 계획발전량컬럼="없음", 예측발전량컬럼="없음",
         저수량저수율컬럼="reservoir_volume", 시간단위="월", 데이터기간="2024-01 ~ 2024-10", 행수=10,
         비고="충주수력 1개소 10행 샘플. 사례기간 미포함, T4 사용 불가"),
    dict(구분="파일", 위치=f"{REPO}/data/processed", 테이블_파일명="daily_dataset.csv / monthly_dataset.csv",
         주요컬럼="댐별 일별·월별 병합 데이터셋(ASOS·SPI 포함)", 날짜컬럼="date / year-month",
         댐발전소명컬럼="dam", 위치컬럼="없음", 실제발전량컬럼="actual_mwh", 계획발전량컬럼="planned_mwh",
         예측발전량컬럼="없음", 저수량저수율컬럼="water_storage_mcm, water_level_elm", 시간단위="일/월",
         데이터기간="K-water 2004-01 ~ 2026-06-06 (원자료 기준)", 행수="파일 참조",
         비고="[중요] 원자료는 2004년부터 보유. DB에는 2021년 이후만 적재됨"),
    dict(구분="파일", 위치=f"{REPO}/output", 테이블_파일명="dam_*_{timestamp}.csv (5종)",
         주요컬럼="DB 적재 전 산출 CSV", 날짜컬럼="동일", 댐발전소명컬럼="dam_name", 위치컬럼="없음",
         실제발전량컬럼="actual_mwh", 계획발전량컬럼="planned_mwh", 예측발전량컬럼="예측 상·하한",
         저수량저수율컬럼="water_storage_mcm", 시간단위="일/월", 데이터기간="2026-06-06 기준 산출분",
         행수="파일 참조", 비고="DB 적재본과 동일 산식"),
]
tbl_df = pd.DataFrame(tbl_rows)

# --------------------------------------------------- 시트 3. 댐발전소_목록
dam_rows = []
for d in DAMS:
    m = DAM_META[d]
    dg = daily[daily.dam_name == d]
    dam_rows.append({
        "댐명": f"{d}댐", "발전소명(추정)": m["stn"], "댐코드": m["code"], "발전소코드": "DB 미보유",
        "시도": m["sido"], "시군구": m["sgg"], "수계/권역": m["river"],
        "위치출처": "configs/config.yaml (DB에는 위치 컬럼 없음)",
        "총저수용량(백만㎥)": m["cap"], "유효저수용량(백만㎥)": m["eff"],
        "저수용량 출처": "외부 문헌 참고값 — 미검증(K-water 제원 대조 필요)",
        "일별기간": f"{dg.generation_date.min():%Y-%m-%d} ~ {dg.generation_date.max():%Y-%m-%d}",
        "일별행수": len(dg),
        "고흥군 관련": "무관", "합천군 관련": "소재지 일치" if d == "합천" else "무관",
        "강릉시 관련": "동일 시도(강원)이나 영서/북한강 — 유역 상이" if d == "소양강" else "무관",
    })
if RDS_OK:
    wp, info = RDS["wp"], RDS["info"]
    for _, r in info.iterrows():
        sid = int(r.STN_ID)
        d = wp[wp.STN_ID == sid]
        pname = d.STN_NAME.iloc[0] if len(d) else "-"
        addr = str(r.ADDRESS)
        allzero = bool(len(d) and (d.POWER == 0).all())
        gn_rel = ("직접 관련 — 강릉시 소재(도암댐 유역변경 방류 수전). "
                  "단 보유구간 발전량 전량 0, 2025년 자료 없음") if sid == 3210 else "무관"
        gh_rel = ("인접 권역이나 고흥군 아님(보성군)" if sid == 3190 else
                  "인접 권역이나 고흥군 아님(정읍시·섬진강)" if sid == 3180 else "무관")
        dam_rows.append({
            "댐명": f"(WT_POWER 발전소) {r.STN_NAME}", "발전소명(추정)": f"WT_POWER 표기: {pname}",
            "댐코드": "-", "발전소코드": sid,
            "시도": addr.split()[0] if addr else "-", "시군구": " ".join(addr.split()[1:2]),
            "수계/권역": "-", "위치출처": "AWS RDS WT_POWER_INFO.ADDRESS (직접 조회)",
            "총저수용량(백만㎥)": "-", "유효저수용량(백만㎥)": "-", "저수용량 출처": "-",
            "일별기간": (f"{d.d.min():%Y-%m-%d} ~ {d.d.max():%Y-%m-%d}" if len(d) else "자료 없음"),
            "일별행수": f"{len(d)} (전량 0값)" if allzero else len(d),
            "고흥군 관련": gh_rel, "합천군 관련": "무관", "강릉시 관련": gn_rel,
        })
dam_df = pd.DataFrame(dam_rows)
dam_df.loc[len(dam_df)] = {
    "댐명": "(고흥군 대응 댐 없음)", "발전소명(추정)": "-", "댐코드": "-", "발전소코드": "-",
    "시도": "전라남도", "시군구": "고흥군", "수계/권역": "영산강·섬진강권 및 해안 소하천",
    "위치출처": "-", "총저수용량(백만㎥)": "-", "유효저수용량(백만㎥)": "-", "저수용량 출처": "-",
    "일별기간": "-", "일별행수": 0, "고흥군 관련": "대응 자료 미보유", "합천군 관련": "-", "강릉시 관련": "-",
}

# --------------------------------------------- 시트 4. 실제발전량_점검
act_rows = []
for d in DAMS:
    dg = daily[daily.dam_name == d]
    gm = gen[gen.dam_name == d]
    rl = real[real.dam_name == d]
    for label, yms in (("2022(남부사례)", range(202201, 202213)), ("2023(남부사례)", range(202301, 202313)),
                       ("2025(강릉사례)", range(202501, 202513))):
        yms = list(yms)
        have_m = sorted(set(gm[gm.ymi.isin(yms)].ymi))
        yr = yms[0] // 100
        dd = dg[dg.generation_date.dt.year == yr]
        exp_days = len(pd.date_range(f"{yr}-01-01", f"{yr}-12-31", freq="D"))
        miss_days = exp_days - len(dd)
        rr = rl[rl.ymi.isin(yms)]
        act_rows.append({
            "댐명": f"{d}댐", "기간": label,
            "실제발전량 컬럼": "actual_mwh", "단위": "MWh",
            "일별 보유일수": f"{len(dd)}/{exp_days}", "일별 결측일수": miss_days,
            "월별 보유월수": f"{len(have_m)}/12",
            "NULL 건수": int(dd.actual_mwh.isna().sum()),
            "0값 건수": int((dd.actual_mwh == 0).sum()),
            "음수 건수": int((dd.actual_mwh < 0).sum()),
            "DB월별값(누계월평균) 평균": round(float(gm[gm.ymi.isin(yms)].actual_mwh.mean()), 1),
            "환산 월발전량 합계(MWh)": round(float(rr.gen_actual_mwh.sum()), 0),
            "환산 월발전량 평균(MWh)": round(float(rr.gen_actual_mwh.mean()), 0),
            "월별 집계 가능": "가능(단, 연누계 차분 환산 필요)",
            "비고": "매년 1월 1일 값은 전년 누계 총량이므로 해당일 실발전량은 복원 불가",
        })
act_df = pd.DataFrame(act_rows)

# ------------------------------------ 시트 5. 계획예측발전량_점검
pj = pred.merge(gen[["dam_name", "ymi", "actual_mwh", "planned_mwh"]], on=["dam_name", "ymi"], how="left")
pj = pj.merge(res[["dam_name", "ymi", "water_storage_mcm"]], on=["dam_name", "ymi"], how="left")
pj["yr"] = pj.ymi // 100
pj["폭"] = pj.pu - pj.pl
pj["중앙"] = (pj.pu + pj.pl) / 2
pj["gen_hit"] = (pj.actual_mwh >= pj.pl) & (pj.actual_mwh <= pj.pu)
pj["sto_hit"] = (pj.water_storage_mcm >= pj.sl) & (pj.water_storage_mcm <= pj.su)

# 레거시 고정밴드 판별: 반폭이 정확히 80.00MWh 이고 구간중앙이 실적과 일치하는 행
pj["is_legacy"] = ((pj["폭"] / 2).round(2) == 80.00) & ((pj["중앙"] - pj.actual_mwh).abs() < 1.0)
LEGACY_LAST = int(pj[pj.is_legacy].ymi.max())      # 레거시 밴드가 마지막으로 쓰인 연월
MODEL_FIRST = int(pj[~pj.is_legacy].ymi.min())     # 모델 산출물이 처음 등장한 연월

pp_rows = []
for yr, g in pj.groupby("yr"):
    n_leg = int(g.is_legacy.sum())
    if n_leg == len(g):
        kind, usable = "레거시 고정밴드(실적 ±80MWh / 저수량 ±40)", \
            "불가 — 실적을 그대로 중심으로 만든 밴드라 오차가 구조적으로 0"
    elif n_leg == 0:
        kind, usable = "모델 산출물(Quantile GBR + conformal)", "가능 — 실적과 독립적으로 산출된 예측구간"
    else:
        leg_months = sorted({int(x) % 100 for x in g[g.is_legacy].ymi})
        kind = f"혼재 — {min(leg_months)}~{max(leg_months)}월 레거시 고정밴드({n_leg}행), 나머지 모델 산출물({len(g)-n_leg}행)"
        usable = "부분 가능 — 모델 산출 구간만 오차계산 가능"
    ev = g.dropna(subset=["actual_mwh"])          # 실적 있는 행만 적중률 계산
    ev_s = g.dropna(subset=["water_storage_mcm"])
    mdl = g[~g.is_legacy]
    pp_rows.append({
        "연도": int(yr), "행수": len(g), "예측 유형": kind,
        "레거시 행수": n_leg, "모델 행수": len(g) - n_leg,
        "계획발전량 존재": "○ dam_monthly_generation.planned_mwh (연누계 월평균)",
        "예측발전량 존재": "○ 상·하한 구간만 (점추정 컬럼 없음)",
        "구간폭 중앙값(MWh)": round(float(g["폭"].median()), 1),
        "구간폭 최대(MWh)": round(float(g["폭"].max()), 1),
        "모델구간 폭 중앙값(MWh)": round(float(mdl["폭"].median()), 1) if len(mdl) else None,
        "구간중앙-실적 절대오차 최대": round(float((g["중앙"] - g.actual_mwh).abs().max()), 2),
        "실적 보유 행수": len(ev),
        "발전량 구간적중률(실적 보유행 기준)": round(float(ev.gen_hit.mean()), 3) if len(ev) else None,
        "저수량 구간적중률(실적 보유행 기준)": round(float(ev_s.sto_hit.mean()), 3) if len(ev_s) else None,
        "실적 대비 오차계산 가능": usable,
        "사례 해당": "남부(2022~2023)" if yr in (2022, 2023) else ("강릉(2025)" if yr == 2025 else "-"),
    })
pp_df = pd.DataFrame(pp_rows)

# --------------------------------------- 시트 6. 저수량저수율_점검
res_rows = []
for d in DAMS:
    rr = res[res.dam_name == d]
    for label, yms in (("2022(남부사례)", SOUTH_YMS[:12]), ("2023(남부사례)", SOUTH_YMS[12:]),
                       ("2025(강릉사례)", GN_YMS)):
        s = rr[rr.ymi.isin(yms)]
        cap = DAM_META[d]["cap"]
        res_rows.append({
            "댐명": f"{d}댐", "기간": label,
            "저수량 컬럼": "water_storage_mcm", "단위": "백만㎥",
            "저수위 컬럼": "water_level_elm (EL.m)",
            "저수율(%) 컬럼": "없음 — DB 미보유",
            "보유월수": f"{len(s)}/{len(yms)}", "NULL": int(s.water_storage_mcm.isna().sum()),
            "월평균 저수량 최소": round(float(s.water_storage_mcm.min()), 1),
            "월평균 저수량 최대": round(float(s.water_storage_mcm.max()), 1),
            "총저수용량 대비 최소(%)": round(float(s.water_storage_mcm.min()) / cap * 100, 1),
            "총저수용량 대비 최대(%)": round(float(s.water_storage_mcm.max()) / cap * 100, 1),
            "발전량과 댐 기준 연결": "○ dam_code + year + month 로 1:1 결합 가능",
            "비고": "총저수용량은 외부 참고값(미검증)이므로 (%) 값은 잠정치",
        })
res_df = pd.DataFrame(res_rows)

# ---------------------------------------- 시트 7. 공간적절성_판정
sp_rows = []
for case, yms in ((SOUTH, SOUTH_YMS), (GN, GN_YMS)):
    for d in DAMS:
        verdict, use, reason = VERDICT[(case, d)]
        m = DAM_META[d]
        rr = real[(real.dam_name == d) & (real.ymi.isin(yms))]
        sp_rows.append({
            "case_id": case, "사례지역": CASE_REGION[case], "대상기간": f"{yms[0]} ~ {yms[-1]}",
            "댐/발전소": f"{d}댐 ({m['stn']})", "댐 위치": f"{m['sido']} {m['sgg']}", "수계/권역": m["river"],
            "사례지역과의 관계": ("소재지 행정구역 일치" if (case == SOUTH and d == "합천")
                          else "동일 시도·상이 유역" if (case == GN and d == "소양강") else "행정구역·유역 모두 상이"),
            "동일 유역/권역 여부": "예" if (case == SOUTH and d == "합천") else "아니오",
            "공간판정": verdict,
            "발전량 변화가 검증에 의미 있는가": (
                "있음 — 사례지역 소재 댐으로 같은 시기 발전량·저수량 변화 비교 가능"
                if verdict == "직접 검증자료" else
                "제한적 — 강원 광역 수자원 상황 서술에 한정" if verdict == "보조자료" else
                "없음 — 유역이 달라 사례 가뭄과 연결 근거가 없음"),
            "억지 포함 시 문제": (
                "-" if verdict == "직접 검증자료" else
                "영동(강릉)과 영서(춘천)는 태백산맥을 경계로 강수·수자원 체계가 달라 동일 지표로 제시하면 오독을 유발한다."
                if verdict == "보조자료" else
                "유역·행정구역이 무관한 댐의 발전량 변화를 사례 가뭄 지표로 제시하면 공간적 근거 없는 연결이 된다."),
            "판정 근거": reason,
            "자료 보유(월)": f"{len(rr)}/{len(yms)}",
            "사용판정": use,
        })
gn_zero = rds_note(lambda: f"{len(RDS['wp'][RDS['wp'].STN_ID == 3210])}행 전부 0", "미확인")
for case, region, target, verdict, use, reason in [
    (SOUTH, "전남 고흥군", "고흥군 대응 댐/발전소", "적용 부적합(자료 미보유)", "미사용",
     "고흥군에 대응하는 다목적댐이 dam_* 5종에 없다. WT_POWER 10개소를 직접 조회한 결과 "
     "전남 소재는 보성강(3190, 전라남도 보성군 득량면) 하나뿐이며 고흥군 소재 발전소는 없다. "
     "보성강·칠보(3180, 전북 정읍시)는 보유 구간이 2023-07-26~12-26뿐이라 "
     "사례 핵심시기(2022 하반기~2023 상반기)를 포함하지 않는다."),
    (GN, "강원 강릉시", "강릉(오봉) / 도암 발전소 — WT_POWER STN_ID 3210", "적용 부적합(자료 부재)", "미사용",
     "공간적으로는 직접 관련된다 — 주소가 '강원도 강릉시 성산면 오봉리'로 강릉시 소재이며, "
     "도암댐 물을 유역변경으로 받아 발전하는 강릉수력이다(WT_POWER_INFO 표기 '강릉', WT_POWER 표기 '도암', "
     "STN_ID 3210으로 동일 발전소 확인). 그러나 ① 보유 구간이 2023-07-26~12-26뿐이고 "
     f"② 그 구간의 POWER 값이 {gn_zero}이며 ③ 2025년 사례기간 자료가 0행이라 검증표에 넣을 수 있는 값이 없다."),
]:
    sp_rows.append({
        "case_id": case, "사례지역": CASE_REGION[case],
        "대상기간": f"{(SOUTH_YMS if case == SOUTH else GN_YMS)[0]} ~ {(SOUTH_YMS if case == SOUTH else GN_YMS)[-1]}",
        "댐/발전소": target, "댐 위치": region, "수계/권역": "-",
        "사례지역과의 관계": "핵심지역 본체", "동일 유역/권역 여부": "예(단 자료 없음)",
        "공간판정": verdict,
        "발전량 변화가 검증에 의미 있는가": "의미 있으나 사례기간 자료가 없어 산출 불가",
        "억지 포함 시 문제": "기간이 다른 자료를 사례 검증표에 넣으면 시점 불일치로 잘못된 비교가 된다.",
        "판정 근거": reason, "자료 보유(월)": "0", "사용판정": use,
    })
sp_df = pd.DataFrame(sp_rows)

# ---------------------------------------- 시트 8. 월별통합가능성
LEGACY_SET = set(map(tuple, pj.loc[pj.is_legacy, ["dam_name", "ymi"]].itertuples(index=False, name=None)))


def build_monthly(case, yms):
    rows = []
    for d in DAMS:
        verdict, use, _ = VERDICT[(case, d)]
        cap = DAM_META[d]["cap"]
        for ym in yms:
            g = gen[(gen.dam_name == d) & (gen.ymi == ym)]
            r = res[(res.dam_name == d) & (res.ymi == ym)]
            p = pred[(pred.dam_name == d) & (pred.ymi == ym)]
            rl = real[(real.dam_name == d) & (real.ymi == ym)]
            yr = ym // 100
            legacy = bool(LEGACY_SET & {(d, ym)})
            act_real = float(rl.gen_actual_mwh.iloc[0]) if len(rl) else None
            plan_real = float(rl.gen_planned_mwh.iloc[0]) if len(rl) else None
            pl = float(p.pl.iloc[0]) if len(p) else None
            pu = float(p.pu.iloc[0]) if len(p) else None
            ctr = (pl + pu) / 2 if pl is not None else None
            sto = float(r.water_storage_mcm.iloc[0]) if len(r) else None
            rows.append({
                "case_id": case, "연월": f"{yr}-{ym % 100:02d}", "지역": CASE_REGION[case],
                "댐/발전소": f"{d}댐", "공간판정": verdict,
                "실제발전량_DB원본_누계월평균(MWh)": float(g.actual_mwh.iloc[0]) if len(g) else None,
                "실제발전량_월환산(MWh)": round(act_real, 1) if act_real is not None else None,
                "계획발전량_DB원본_누계월평균(MWh)": float(g.planned_mwh.iloc[0]) if len(g) else None,
                "계획발전량_월환산(MWh)": round(plan_real, 1) if plan_real is not None else None,
                "예측발전량_하한(MWh)": pl, "예측발전량_상한(MWh)": pu,
                "예측유형": "레거시 고정밴드(모델 아님)" if legacy else "모델 산출물",
                "발전량오차_실적-예측중앙(MWh)": (round(float(g.actual_mwh.iloc[0]) - ctr, 1)
                                       if (len(g) and ctr is not None) else None),
                "발전량오차 사용가능": "불가(구조적 0)" if legacy else "가능",
                "저수량(백만㎥)": sto,
                "저수위(EL.m)": float(r.water_level_elm.iloc[0]) if len(r) else None,
                "저수율(%)_잠정": round(sto / cap * 100, 1) if sto is not None else None,
                "사용판정": use,
                "비고": ("사례지역 소재 댐" if verdict == "직접 검증자료" else
                       "강원 광역 보조지표로만 참고" if verdict == "보조자료" else "공간 연관성 낮아 검증표 제외"),
            })
    return rows


mon_df = pd.DataFrame(build_monthly(SOUTH, SOUTH_YMS) + build_monthly(GN, GN_YMS))

# ---------------------------------------- 시트 9. 미비점판정
gen_days_missing = int(len(pd.date_range(daily.generation_date.min(), daily.generation_date.max(), freq="D"))
                       - daily.generation_date.nunique())
lack = [
    ("수력발전량 테이블 확인", "완료",
     f"로컬·RDS 양쪽 조회 완료. dam_* 5종(로컬 {len(daily)}/{len(gen)}/{len(res)}/{len(pred)}/{len(comp)}행) + "
     f"RDS 전용 WT_POWER·WT_POWER_INFO + drought_impact_*_dam_operation 4종 확인",
     "RDS 전용 테이블은 로컬에 복제되어 있지 않아 파이프라인이 인지하지 못한다",
     "T4에서 쓸 테이블 목록을 확정하고, RDS 전용 4종의 출처·컬럼 의미를 데이터 제공자에게 확인"),
    ("대상 댐/발전소 목록", "완료",
     "dam_* = 소양강·충주·대청·합천 4개(코드 1012110/1003110/3008110/2015110). "
     "WT_POWER_INFO = 화천·춘천·안흥(강림)·의암·청평·팔당·괴산·섬진강(칠보)·보성강·강릉(도암) 10개소. "
     "drought_impact_*_dam_operation = 발전소코드 1070/1080/1090/1110",
     "dam_* 에는 발전소명·발전소코드 컬럼이 없고, 세 계열의 발전소 코드 체계가 서로 다르다"
     "(dam_code 7자리 / STN_ID 4자리 / elcpwstnCd 4자리)",
     "세 체계를 잇는 발전소 매핑표 작성(합천댐 2015110 ↔ 합천수력 1110 등)"),
    ("댐/발전소 위치 확인", "진행 중",
     "dam_* 5종 전 56개 컬럼에 주소/좌표/시도 컬럼 없음. 위치는 configs/config.yaml의 province/city로만 확인",
     "DB만으로는 공간판정 불가. 위치정보가 코드 저장소에만 있음",
     "댐 위치(시도·시군구·수계·좌표)를 마스터 테이블 또는 T4 지역마스터에 명시적으로 결합"),
    ("실제 발전량", "진행 중",
     f"actual_mwh 일별 {daily.generation_date.min():%Y-%m-%d}~{daily.generation_date.max():%Y-%m-%d}, "
     "NULL 0건, 음수 0건. 2022·2023·2025 전 기간 보유",
     "[중대] 값이 '연간 누계'라 월별 테이블(누계의 월평균)을 월 발전량으로 쓰면 안 됨. "
     "매년 1월 1일은 전년 누계 총량 이상치",
     "월말 누계 차분으로 실제 월별 발전량 환산 후 T4 지표로 사용(본 워크북 '월별통합가능성' 시트에 환산값 제공)"),
    ("계획 발전량", "완료",
     "planned_mwh가 일별·월별 모두 존재. NULL 0건",
     "실적과 동일하게 연누계 성격이므로 동일 환산 필요",
     "환산 후 계획 대비 실적 비교 지표 산출"),
    ("예측 발전량", "진행 중",
     f"dam_monthly_predictions 2022-01~2026-08, 댐별 56개월 결측 없음. "
     f"{LEGACY_LAST//100}-{LEGACY_LAST%100:02d}까지 구간폭 정확히 ±80MWh 고정이고 구간중앙-실적 오차 최대 0.05MWh, "
     f"{MODEL_FIRST//100}-{MODEL_FIRST%100:02d}부터 모델 산출물",
     "[중대] 2022-01~2024-10 값은 실적으로 역산한 고정밴드라 '예측'이 아님 → 남부 사례기간 예측검증 불가. "
     "점추정(중앙값) 컬럼도 없음",
     "남부 사례는 예측 지표 제외. 필요 시 backtest 명령으로 2022~2023 시점별 재학습 예측을 새로 생성"),
    ("저수량/저수율", "진행 중",
     f"water_storage_mcm·water_level_elm {rng(res)} 결측 0건, 댐별 66개월",
     "저수율(%) 컬럼 없음. 총저수용량 제원이 DB에 없어 정규화 불가",
     "K-water 댐 제원(총/유효 저수용량) 확보 후 저수율 산정, 또는 저수량 원값으로만 제시"),
    ("월별 집계 가능성", "완료",
     "일별→월평균 재집계가 dam_monthly_generation과 오차 0.0으로 일치(합천 2022-01~2023-12 검증). "
     "dam_code+year+month로 5종 테이블 결합 가능",
     "집계 규칙이 '월평균'이라 월 합계와 다름. 연누계 특성과 겹쳐 해석 혼동 위험",
     "T4 산출물에 집계 규칙(월평균/누계차분)을 컬럼명과 메타에 명시"),
    ("2022~2023 남부 사례 적용성", "진행 중",
     "4개 댐 모두 24/24개월 발전량·저수량·예측 보유. 합천댐 월환산 발전량은 2023 상반기 전년동월 대비 -42~-63% 범위",
     "고흥군 대응 댐 없음. 예측값은 레거시 고정밴드라 사용 불가",
     "합천댐만 직접 검증자료로 채택하고 고흥군은 수력발전량 지표 미적용으로 명시"),
    ("2025 강릉 사례 적용성", "막힘",
     "4개 댐 모두 12/12개월 보유하나 강릉 유역 댐은 0개. WT_POWER에서 강릉시 소재 발전소(STN_ID 3210, "
     "강원도 강릉시 성산면 오봉리)를 찾았으나 보유 구간이 2023-07-26~12-26뿐이고 해당 92행이 전부 POWER=0, "
     "2025년 자료 0행",
     "강릉 사례에 공간적으로 연결되는 수력발전 자료가 사례기간(2025)에 존재하지 않는다. "
     "보유 구간조차 발전 실적이 0이라 대체 기간 비교도 불가",
     "① 강릉 사례는 수력발전량 지표 제외로 확정하거나, ② 강릉수력(도암) 2025년 발전실적을 "
     "외부(한국수력원자력·EPSIS)에서 확보할지 의사결정. 확보 전까지는 T4 검증표에서 제외"),
    ("공간 적절성 판정", "완료",
     "사례×댐 8조합 + 사례지역 대응자료 2건 판정 완료('공간적절성_판정' 시트). "
     "WT_POWER 10개소 주소를 직접 조회해 고흥군 소재 0개소, 강릉시 소재 1개소(3210) 확인",
     "선행조사의 미해소 항목이던 WT_POWER↔WT_POWER_INFO 명칭 불일치는 해소됨"
     "(STN_ID 조인 시 10/10 매칭, 주소 9/10 일치, 명칭만 3건 상이)",
     "매칭 키를 STN_NAME이 아닌 STN_ID로 고정해 문서화"),
    ("DB 적재 여부", "진행 중",
     "5종 모두 로컬·RDS에 적재됨. 표본 대조 결과 값은 양쪽 동일"
     "(합천 2022-06 42,665.93 / 2023-06 20,113.03 / 2025-06 63,809.60 일치)",
     "[중대] RDS의 dam_* 4종에 인덱스가 0개(PK 없음)라 적재가 멱등하지 않고 완전 동일 행이 반복 저장됨. "
     "dam_daily_generation 20,892행 중 고유 7,660(초과 13,232행), monthly_generation 696/256, "
     "reservoir 456/264, comparison 408/160. predictions만 중복 없음. 값 충돌은 없음",
     "RDS에 UNIQUE 제약(dam_code+generation_date / dam_code+year+month) 추가 후 중복 정리. "
     "그 전까지 T4 집계는 반드시 SELECT DISTINCT로 중복 제거 후 수행"),
    ("보고용 산출물 존재 여부", "완료",
     f"본 워크북 {OUT_PATH} 생성. 선행 산출물로 {SURVEY_MD}, "
     f"{REPO}/output 의 dam_* CSV, models/metrics.json 존재",
     "T4 전용 월별 검증표가 그동안 없었음",
     "본 워크북 '월별통합가능성' 시트를 T4 표준 입력으로 채택"),
]
lack_df = pd.DataFrame(lack, columns=["항목", "상태", "확인 근거", "미비점", "필요한 작업"])

# ------------------------------------- 시트 10. 추출쿼리_파일경로
qry = [
    ("테이블 목록", "SHOW TABLES; / SELECT table_schema, table_name FROM information_schema.tables "
                "WHERE table_name LIKE '%POWER%' OR table_name LIKE '%dam%' OR table_name LIKE '%hydro%';",
     "ACTUAL_DRGHT 27개, v2 21개, v3 26개, v4 34개, drought 7개 테이블 확인"),
    ("컬럼 구조", "SHOW FULL COLUMNS FROM dam_daily_generation; (5종 반복)", "컬럼명·타입·주석 확인"),
    ("댐 목록/기간", "SELECT dam_name, dam_code, COUNT(*), MIN(generation_date), MAX(generation_date) "
                "FROM dam_daily_generation GROUP BY dam_name, dam_code;", "4개 댐 × 1,915일"),
    ("월별 커버리지", "SELECT dam_name, CAST(year AS UNSIGNED)*100+CAST(month AS UNSIGNED) ymi FROM "
                 "dam_monthly_generation; (reservoir/predictions/comparison 동일)", "사례기간 결측 확인"),
    ("연누계 확인", "SELECT generation_date, planned_mwh, actual_mwh FROM dam_daily_generation "
                "WHERE dam_name='합천' AND generation_date BETWEEN '2022-12-28' AND '2023-01-05' ORDER BY 1;",
     "2023-01-01=88,388(2022 누계) → 2023-01-02=171 로 리셋되는 것 확인"),
    ("월평균 일치 검증", "일별 actual_mwh의 월평균 vs dam_monthly_generation.actual_mwh 비교",
     "합천 2022-01~2023-12 최대 절대오차 0.0"),
    ("예측 밴드 규칙", "dam_monthly_predictions ⨝ dam_monthly_generation 후 (pu-pl)/2 및 (pu+pl)/2-actual 산출",
     "2022~2024 반폭 정확히 80.00, 중앙-실적 오차 최대 0.05 → 실적 역산 고정밴드"),
    ("결측일 산출", "date_range(min,max) 대비 보유 generation_date 차집합",
     "2021-12-03, 2022-05-27, 2024-05-31, 2024-10-28~12-31(65일)"),
    ("파일 경로", f"{REPO}/configs/config.yaml", "댐 코드·시도·시군구·ASOS 지점 매핑 (유일한 위치 출처)"),
    ("파일 경로", f"{REPO}/data/processed/monthly_dataset.csv", "K-water 2004~ 월별 데이터셋"),
    ("파일 경로", f"{REPO}/output/dam_*_202606071702.csv", "DB 적재 직전 산출 CSV 5종"),
    ("파일 경로", f"{REPO}/models/metrics.json, models/conformal.json", "모델 성능·conformal 보정계수"),
    ("RDS 접속", "ssh -i ~/Desktop/aws-key/icuh.cer -f -N -L 13306:icuh-rds.cd7bwwfid5u3.ap-northeast-2."
                "rds.amazonaws.com:3306 ubuntu@ec2-3-37-105-126.ap-northeast-2.compute.amazonaws.com",
     "EC2 경유 SSH 터널 후 127.0.0.1:13306 으로 ACTUAL_DRGHT 조회 (읽기 전용)"),
    ("RDS 중복 확인", "SELECT COUNT(*) FROM dam_daily_generation; vs "
                  "SELECT COUNT(*) FROM (SELECT DISTINCT dam_code, generation_date FROM dam_daily_generation) x; "
                  "/ SHOW INDEX FROM dam_daily_generation;",
     "20,892 vs 7,660, 인덱스 0개 → PK 없이 중복 적재됨"),
    ("WT_POWER 매칭", "SELECT i.STN_ID, i.STN_NAME, p.STN_NAME, i.ADDRESS FROM WT_POWER_INFO i "
                  "LEFT JOIN WT_POWER p ON p.STN_ID=i.STN_ID GROUP BY 1,2,3,4;",
     "10/10 매칭, 명칭 3건 상이(3130·3180·3210), 주소 9/10 일치"),
    ("강릉 발전소 확인", "SELECT * FROM WT_POWER WHERE STN_ID=3210 ORDER BY CRT_YMD;",
     "강원도 강릉시 성산면 오봉리, 2023-07-26~12-26 92일, POWER 전량 0, 2025년 0행"),
    ("r2/s2 의미 검증", "drought_impact_*_dam_operation 의 r2·s2·t2 를 K-water 누계차분 월발전량 및 "
                   "월평균 저수량과 상관분석",
     "r2 = 월 발전량(상관 +0.996~0.999), s2 = 의미 미상"),
    ("파일 경로", SURVEY_MD, "선행 RDS 현황조사(2026-08-05) — 본 점검에서 WT_POWER 관련 내용 갱신·정정"),
    ("파일 경로", f"{OUT_DIR}/실측가뭄_T3_뉴스영향데이터_정리_v1.xlsx", "T3 산출물(case_id 체계 정합)"),
    ("점검 스크립트", "/private/tmp/.../scratchpad/q1.py ~ q4.py, build_t4_hydro_workbook.py",
     "본 점검에 사용한 조회·집계 스크립트(임시 디렉터리)"),
]
qry_df = pd.DataFrame(qry, columns=["구분", "쿼리 / 경로", "확인 결과"])

# ---------------------------------------- 시트 11. 품질점검
qc = [
    ("일자 결측", "dam_daily_generation", "주의",
     f"전 댐 공통 {gen_days_missing}일 결측: 2021-12-03, 2022-05-27, 2024-05-31, 2024-10-28~2024-12-31(65일)",
     "남부 사례기간은 2022-05-27 1일만 결측(729/730), 강릉 사례기간은 결측 0(365/365)"),
    ("중복", "dam_daily_generation", "양호", "dam_name+generation_date 중복 0건", "-"),
    ("중복", "월별 3종", "양호", "dam_name+year+month 중복 0건(generation/reservoir/predictions)", "-"),
    ("결측값(NULL)", "전 테이블", "양호",
     "actual_mwh·planned_mwh·water_storage_mcm·water_level_elm 결측 0건", "-"),
    ("0값/음수", "dam_daily_generation", "양호",
     "actual_mwh=0 은 2021년 각 댐 1건씩(총 4건), 음수 0건", "2021년은 사례기간 밖"),
    ("값 의미 불일치", "actual_mwh / planned_mwh", "중대",
     "연간 누계 값. 매년 1월 1일에 전년 누계 총량이 기록되고 1월 2일부터 리셋됨",
     "월별 테이블은 이 누계의 '월평균'이므로 월 발전량으로 직접 해석 불가 → 누계 차분 환산 필요"),
    ("값 의미 불일치", "dam_monthly_predictions 2022-01~2024-10", "중대",
     "구간 반폭이 정확히 80.00MWh 고정, 구간중앙과 실적의 최대 오차 0.05MWh (해당 구간 전 행)",
     "모델 예측이 아니라 실적에서 역산한 고정밴드 → 예측정확도 검증 불가. 2024-11부터는 모델 산출물"),
    ("구간 폭 과대", "dam_monthly_predictions 2025", "주의",
     "2025-01 구간: 대청 0~156,803 / 소양강 0~338,403 / 합천 3,905~182,122 MWh",
     "conformal 보정으로 구간이 과도하게 넓어져 해당 월 예측의 실용성이 낮음"),
    ("구간 적중률", "dam_monthly_predictions 2025", "양호",
     "발전량 70.8%, 저수량 95.8%", "발전량은 목표 80% 미달(알려진 한계)"),
    ("월별 결측", "dam_monthly_generation", "주의",
     f"2021-01~2026-06 중 댐별 64/66개월 보유. 2024-11·2024-12 전 댐 결측 "
     f"(dam_monthly_reservoir_status는 66/66으로 결측 없음)",
     "사례기간(2022~2023, 2025)에는 영향 없으나 2025-11·2025-12의 전년동월 대비 산출이 불가"),
    ("comparison 결측", "dam_monthly_comparison", "주의",
     "2022년 전체 없음(2023-01부터 시작), 2024-11·2024-12 결측",
     "2024-10-28~12-31 수집 공백의 영향. 2025-11·2025-12의 전년동월 대비 계산 불가"),
    ("위치정보 부재", "dam_* 5종", "중대",
     "56개 컬럼 중 주소·좌표·시도·시군구 컬럼 0개",
     "공간판정을 DB만으로 수행할 수 없어 configs/config.yaml에 의존"),
    ("저수율 부재", "dam_monthly_reservoir_status", "주의",
     "저수량(백만㎥)·저수위(EL.m)만 보유, 저수율(%) 없음",
     "총저수용량 제원 확보 후 산정 필요. 본 워크북의 저수율은 외부 참고값 기반 잠정치"),
    ("중복 적재", "AWS RDS dam_* 4종", "중대",
     rds_note(lambda: " / ".join(
         f"{t.replace('dam_','')} 총 {v['total']}행·고유 {v['unique']}·초과 {v['excess']}"
         for t, v in RDS["dup"].items()) + " | 인덱스 0개(PK 없음)", "RDS 미접속으로 미확인"),
     "PK/UNIQUE가 없어 적재가 멱등하지 않다. 값 충돌은 없고 로컬과 값이 일치하므로 "
     "SELECT DISTINCT로 제거 가능하나, 제거하지 않으면 월평균·합계가 왜곡된다"),
    ("컬럼 주석 오류", "drought_impact_*_dam_operation", "중대",
     "주석상 '저수량(ton)'인 r2가 4개 댐 모두 월 발전량과 상관 +0.996~+0.999(중앙비율 0.995~1.000), "
     "저수량과는 -0.22~+0.36. 주석상 '발전량(Mwh)'인 s2는 발전량 상관 -0.19~+0.47로 의미 미상",
     "컬럼명만 믿고 s2를 발전량으로 쓰면 잘못된 값이 들어간다. 출처 확인 전 사용 보류"),
    ("WT_POWER 품질", "AWS RDS WT_POWER", "주의",
     rds_note(lambda: f"총 {RDS['wp_total']}행 중 고유 {len(RDS['wp'])}행(중복 존재). "
                      f"기간 2023-07-26~12-26 중 고유 {RDS['wp'].d.nunique()}일만 존재(결측 62일). "
                      f"POWER 단위 미표기", "RDS 미접속으로 미확인"),
     "두 사례기간을 모두 벗어나 T4 검증표에 직접 사용 불가"),
    ("원자료 범위 대비 적재 범위", "K-water 원자료 vs DB", "주의",
     "원자료는 2004-01부터 보유(data/processed)하나 DB에는 2021-01부터만 적재",
     "장기 평년 대비 분석이 필요하면 2004~2020 구간 추가 적재 검토"),
]
qc_df = pd.DataFrame(qc, columns=["점검 항목", "대상", "판정", "확인 내용", "T4 영향"])

# ------------------------------------------------------------------ 쓰기
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUS_FILL = {
    "완료": PatternFill("solid", fgColor="D6E9C6"), "진행 중": PatternFill("solid", fgColor="FDF3C6"),
    "미착수": PatternFill("solid", fgColor="E8E8E8"), "막힘": PatternFill("solid", fgColor="F8CBCB"),
    "중대": PatternFill("solid", fgColor="F8CBCB"), "주의": PatternFill("solid", fgColor="FDF3C6"),
    "양호": PatternFill("solid", fgColor="D6E9C6"), "확인 필요": PatternFill("solid", fgColor="DCE6F1"),
    "직접 검증자료": PatternFill("solid", fgColor="D6E9C6"),
    "보조자료": PatternFill("solid", fgColor="FDF3C6"),
    "적용 부적합": PatternFill("solid", fgColor="F8CBCB"),
    "적용 부적합(자료 미보유)": PatternFill("solid", fgColor="F8CBCB"),
}
WIDTHS = {"README": [26, 118]}

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as xw:
    pd.DataFrame(readme).to_excel(xw, sheet_name="README", index=False, header=False)
    sheets = [
        ("테이블_파일목록", tbl_df), ("댐발전소_목록", dam_df), ("실제발전량_점검", act_df),
        ("계획예측발전량_점검", pp_df), ("저수량저수율_점검", res_df), ("공간적절성_판정", sp_df),
        ("월별통합가능성", mon_df), ("미비점판정", lack_df), ("추출쿼리_파일경로", qry_df), ("품질점검", qc_df),
    ]
    for name, df in sheets:
        df.to_excel(xw, sheet_name=name, index=False)

    wb = xw.book
    for ws in wb.worksheets:
        is_readme = ws.title == "README"
        if is_readme:
            ws["A1"].font = TITLE_FONT
            for w, col in zip(WIDTHS["README"], "AB"):
                ws.column_dimensions[col].width = w
            for row in ws.iter_rows():
                for c in row:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.freeze_panes = "A2"
            continue

        for c in ws[1]:
            c.fill, c.font = HDR_FILL, HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 32

        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            vals = [len(str(c.value)) if c.value is not None else 0 for c in col]
            width = min(max(12, max(vals) + 2), 60)
            ws.column_dimensions[get_column_letter(i)].width = width

        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.border = BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(c.value, str) and c.value in STATUS_FILL:
                    c.fill = STATUS_FILL[c.value]
                    c.font = Font(bold=True, size=10)
                elif isinstance(c.value, float):
                    c.number_format = "#,##0.0"

print("생성 완료:", OUT_PATH)
print("시트:", ["README"] + [n for n, _ in sheets])
print("월별통합가능성 행수:", len(mon_df))
print("공간적절성 행수:", len(sp_df))
