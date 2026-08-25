# -*- coding: utf-8 -*-
"""Build drought-news rows from a reviewed URL candidate workbook.

This script intentionally does not write to the source DB. It reads the
previously reviewed URL list, fetches article pages when possible, and writes a
JSON intermediate that the spreadsheet builder turns into an xlsx workbook.
"""

from __future__ import annotations

import html
import json
import re
import time
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


CANDIDATE_XLSX = Path("/Users/jeongseok/Desktop/9월실측가뭄워크샵/실측가뭄_고흥합천_가뭄뉴스_재수집_v1.xlsx")
OUTPUT_JSON = Path("/Users/jeongseok/Desktop/workspace_intelliJ/icuh-platform-api/outputs/db_check_work/goheung_hapcheon_news_result_rows.json")

SI_CODES = {
    "전남 고흥군": ("전라남도", "고흥군", 46770),
    "경남 합천군": ("경상남도", "합천군", 48890),
    "경남": ("경상남도", "합천군", 48890),
}

STOPWORDS = {
    "기자",
    "뉴스",
    "기사",
    "지난",
    "이번",
    "대한",
    "위해",
    "관련",
    "통해",
    "지역",
    "가운데",
    "현재",
    "따르면",
    "것으로",
    "있다",
    "했다",
    "한다",
    "등",
}


class MetadataParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_title = False
        self.title_parts: list[str] = []
        self.meta: dict[str, str] = {}
        self.body_chunks: list[str] = []
        self.capture_stack: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attr = {k.lower(): (v or "") for k, v in attrs}
        if tag.lower() == "title":
            self.in_title = True
        if tag.lower() == "meta":
            key = (attr.get("property") or attr.get("name") or "").lower()
            if key in {"og:title", "twitter:title", "description", "og:description", "twitter:description", "article:published_time"}:
                self.meta[key] = attr.get("content", "")
        if tag.lower() in {"article", "p", "div"}:
            marker = " ".join([attr.get("id", ""), attr.get("class", "")]).lower()
            if tag.lower() == "article" or any(
                token in marker
                for token in [
                    "article",
                    "news_view",
                    "newsct_article",
                    "article_body",
                    "news_body",
                    "view_cont",
                    "article-view-content",
                    "articlebody",
                    "art_txt",
                    "content",
                    "article_txt",
                ]
            ):
                self.capture_stack.append(tag.lower())

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "title":
            self.in_title = False
        if self.capture_stack and tag.lower() == self.capture_stack[-1]:
            self.capture_stack.pop()

    def handle_data(self, data: str) -> None:
        text = re.sub(r"\s+", " ", data).strip()
        if not text:
            return
        if self.in_title:
            self.title_parts.append(text)
        if self.capture_stack and len(text) >= 20:
            self.body_chunks.append(text)


def clean_text(text: str | None) -> str:
    if not text:
        return ""
    text = html.unescape(str(text))
    text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def first_sentence_blob(parts: list[str], limit: int = 900) -> str:
    seen = []
    for part in parts:
        c = clean_text(part)
        if len(c) < 20:
            continue
        if c in seen:
            continue
        seen.append(c)
        if sum(len(x) for x in seen) > limit:
            break
    return clean_text(" ".join(seen))[:limit]


def fetch_article(url: str) -> dict[str, str | int | bool]:
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    }
    req = Request(url, headers=headers)
    try:
        with urlopen(req, timeout=12) as response:
            status = getattr(response, "status", 200)
            raw = response.read(2_000_000)
            content_type = response.headers.get("Content-Type", "")
    except HTTPError as exc:
        return {"ok": False, "status": exc.code, "error": f"HTTPError: {exc.code}", "title": "", "body": "", "date": ""}
    except URLError as exc:
        return {"ok": False, "status": 0, "error": f"URLError: {exc.reason}", "title": "", "body": "", "date": ""}
    except Exception as exc:
        return {"ok": False, "status": 0, "error": f"{type(exc).__name__}: {exc}", "title": "", "body": "", "date": ""}

    encoding = "utf-8"
    match = re.search(r"charset=([\w.-]+)", content_type, flags=re.I)
    if match:
        encoding = match.group(1)
    text = raw.decode(encoding, errors="replace")

    parser = MetadataParser()
    try:
        parser.feed(text)
    except Exception:
        pass

    title = clean_text(parser.meta.get("og:title") or parser.meta.get("twitter:title") or " ".join(parser.title_parts))
    title = re.sub(r"\s*[-|:]\s*(연합뉴스|뉴스핌|농민신문|한겨레|채널A|국제신문|매일신문|아시아경제|네이트뉴스|다음뉴스).*$", "", title)

    body = clean_text(parser.meta.get("og:description") or parser.meta.get("description"))
    if not body:
        body = first_sentence_blob(parser.body_chunks)

    published = parser.meta.get("article:published_time", "")
    return {
        "ok": True,
        "status": int(status),
        "error": "",
        "title": title,
        "body": body,
        "date": published[:10] if published else "",
        "bytes": len(raw),
    }


def hangul_terms(text: str) -> list[str]:
    cleaned = re.sub(r"[^ㄱ-ㅎㅏ-ㅣ가-힣 ]", " ", text or "")
    terms = []
    for token in re.findall(r"[가-힣]{2,}", cleaned):
        if token in STOPWORDS:
            continue
        terms.append(token)
    deduped = []
    for token in terms:
        if token not in deduped:
            deduped.append(token)
    return deduped[:40]


def list_string(values: list[str | int]) -> str:
    return "[" + ", ".join(repr(v) for v in values) + "]"


def damage_terms(summary: str) -> list[str]:
    raw = re.split(r"[,/·]| 및 |과 |와 ", summary or "")
    terms = []
    for item in raw:
        item = clean_text(item)
        item = re.sub(r"(우려|대응|지원|추진|확보|총력|개선)$", "", item).strip()
        if len(item) >= 2:
            terms.append(item)
    deduped = []
    for term in terms:
        if term not in deduped:
            deduped.append(term)
    return deduped[:8]


def looks_broken(text: str) -> bool:
    if not text:
        return False
    return "�" in text or text.count("����") > 0


def looks_junk(text: str) -> bool:
    lowered = (text or "").lower()
    junk_tokens = [
        "document.",
        "function(",
        "$('.",
        ".slick",
        "window.",
        "var ",
        "뉴스 구독",
        "로그인",
        "회원가입",
    ]
    return any(token in lowered for token in junk_tokens)


def read_candidates() -> list[dict[str, object]]:
    wb = load_workbook(CANDIDATE_XLSX, data_only=True)
    ws = wb["후보기사"]
    headers = [cell.value for cell in ws[3]]
    rows = []
    for values in ws.iter_rows(min_row=4, values_only=True):
        if not values[0]:
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def main() -> None:
    candidates = read_candidates()
    result_rows = []
    status_rows = []

    for idx, row in enumerate(candidates, start=1):
        url = str(row.get("URL") or "").strip()
        fetched = fetch_article(url) if url else {"ok": False, "status": 0, "error": "missing url", "title": "", "body": "", "date": ""}
        time.sleep(0.4)

        date_value = row.get("기사일자")
        if isinstance(date_value, datetime):
            date_text = date_value.strftime("%Y-%m-%d")
        else:
            date_text = str(date_value)[:10]
        if fetched.get("date"):
            date_text = str(fetched["date"])

        title = clean_text(str(fetched.get("title") or ""))
        if not title or looks_broken(title):
            title = clean_text(str(row.get("뉴스제목") or ""))
        body = clean_text(str(fetched.get("body") or ""))
        fallback_used = False
        if len(body) < 30 or looks_broken(body) or looks_junk(body):
            body = clean_text(f"{row.get('뉴스제목') or ''} {row.get('영향요약') or ''} {row.get('검증활용근거') or ''}")
            fallback_used = True
        body = body[:500]

        province, sigungu, sigungu_code = SI_CODES.get(str(row.get("표준지역명") or ""), ("", str(row.get("상세지역") or ""), ""))
        if not province and "합천" in str(row.get("상세지역") or ""):
            province, sigungu, sigungu_code = SI_CODES["경남 합천군"]
        if not province and "고흥" in str(row.get("상세지역") or ""):
            province, sigungu, sigungu_code = SI_CODES["전남 고흥군"]

        impact = clean_text(str(row.get("영향분야") or ""))
        summary = clean_text(str(row.get("영향요약") or ""))

        result_rows.append(
            {
                "일자": date_text,
                "광역시도": list_string([province] if province else []),
                "시군구": list_string([sigungu] if sigungu else []),
                "뉴스링크": url,
                "뉴스제목": title,
                "뉴스제목_명사": list_string(hangul_terms(title)),
                "기사본문": body,
                "기사본문_명사": list_string(hangul_terms(body)),
                "기상요소": "가뭄",
                "시군구코드": list_string([sigungu_code] if sigungu_code else []),
                "영향구분": impact,
                "피해상세": list_string(damage_terms(summary)),
            }
        )
        status_rows.append(
            {
                "article_id": row.get("article_id"),
                "URL": url,
                "수집성공": bool(fetched.get("ok")),
                "HTTP상태": fetched.get("status"),
                "본문fallback사용": fallback_used,
                "본문길이": len(body),
                "오류": fetched.get("error") or "",
                "원래제목": row.get("뉴스제목"),
                "최종제목": title,
                "사용판정": row.get("사용판정"),
                "연결수준": row.get("연결수준"),
            }
        )

        print(f"[{idx}/{len(candidates)}] {fetched.get('status')} fallback={fallback_used} {title[:40]}")

    OUTPUT_JSON.write_text(
        json.dumps({"rows": result_rows, "status": status_rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[INFO] saved {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
